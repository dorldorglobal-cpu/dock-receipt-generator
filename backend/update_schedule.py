#!/usr/bin/env python3
"""
update_schedule.py
Parse Sallaum and ACL PDFs and update master-schedule.xlsx.

Usage:
    python update_schedule.py <sallaum_pdf> <acl_pdf> [master_xlsx]
"""

import sys
import pdfplumber
from datetime import datetime
from openpyxl import load_workbook
import os

DEFAULT_MASTER = os.path.join(os.path.dirname(__file__), "saved-schedules", "master-schedule.xlsx")
YEAR = 2026

SALLAUM_POL_MAP = {
    "Freeport": "FREEPORT",
    "Jacksonville": "JACKSONVILLE",
    "Baltimore Tradepoint": "BALTIMORE TRADEPOINT",
    "Baltimore South Locus Point": "BALTIMORE SOUTH LOCUST",
    "Brunswick GA": "BRUNSWICK GA",
    "NORAD Davisville": "DAVISVILLE",
}

ACL_POL_MAP = {
    "Freeport (TX)": "FREEPORT",
    "Jacksonville": "JACKSONVILLE",
    "Baltimore": "BALTIMORE",
    "Wilmington": "WILMINGTON",
    "Providence": "PROVIDENCE",
}

SALLAUM_POD_MAP = {
    "Cotonou": "COTONOU",
    "Lome": "LOME",
    "Lagos": "LAGOS",
    "Durban (South Africa)": "DURBAN",
}

ACL_POD_MAP = {
    "Dakar": "DAKAR",
    "Lagos": "LAGOS",
    "Cotonou": "COTONOU",
    "Lome": "LOME",
    "Tema": "TEMA",
}


def parse_sallaum_date(s):
    if not s or s.strip() in ("", "N/A", "NA"):
        return "NA"
    try:
        d = datetime.strptime(f"{s.strip()}-{YEAR}", "%d-%b-%Y")
        return d
    except ValueError:
        return "NA"


def parse_acl_date(s):
    if not s or s.strip() in ("", "N/A", "NA", "t/s via Dakar"):
        return "NA"
    try:
        d = datetime.strptime(f"{s.strip()}/{YEAR}", "%m/%d/%Y")
        return d
    except ValueError:
        return "NA"


def parse_sallaum(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        table = pdf.pages[0].extract_tables()[0]

    # Vessel names at odd columns 1,3,5,...,13 in rows 0 and 1
    vessels = []
    for col in range(1, 15, 2):
        name = table[0][col]
        voyage = table[1][col]
        if name and name.strip():
            vessels.append((name.strip().upper(), voyage.strip() if voyage else ""))

    # Find header/POD row boundaries
    header_row = next(i for i, r in enumerate(table) if r[0] == "POL")
    pod_row = next(i for i, r in enumerate(table) if r[0] == "POD")

    # Parse POL cutoff + sail dates per (vessel_idx, pol)
    pol_data = {}  # (v_idx, pol) -> (cutoff, sail)
    for row in table[header_row + 1 : pod_row]:
        pol_name = row[0]
        if not pol_name or pol_name not in SALLAUM_POL_MAP:
            continue
        pol = SALLAUM_POL_MAP[pol_name]
        for v_idx in range(len(vessels)):
            cutoff = parse_sallaum_date(row[2 * v_idx + 1])
            sail = parse_sallaum_date(row[2 * v_idx + 2])
            pol_data[(v_idx, pol)] = (cutoff, sail)

    # Parse POD arrival dates per (vessel_idx, pod)
    rows = []
    for row in table[pod_row + 1 :]:
        pod_name = row[0]
        if not pod_name or pod_name not in SALLAUM_POD_MAP:
            continue
        pod = SALLAUM_POD_MAP[pod_name]
        for v_idx, (vessel, voyage) in enumerate(vessels):
            arrival = parse_sallaum_date(row[2 * v_idx + 2])
            if arrival == "NA":
                continue
            for pol in SALLAUM_POL_MAP.values():
                cutoff, sail = pol_data.get((v_idx, pol), ("NA", "NA"))
                rows.append((vessel, voyage, pol, pod, cutoff, sail, arrival))

    return rows


def parse_acl_sub_table(table, start_row, end_row):
    """Parse one ACL sub-table (rows start_row..end_row)."""
    # Vessels at odd columns 1,3,5,...,11 in first two rows of sub-table
    vessels = []
    for col in range(1, 13, 2):
        name = table[start_row][col]
        voyage = table[start_row + 1][col]
        if name and name.strip():
            vessels.append((name.strip().upper(), voyage.strip() if voyage else ""))

    # Header row is start_row+2 (POL / ETA / Latest Delivery ...)
    pol_header = start_row + 2
    # Find POD marker row
    pod_row = next(
        i for i in range(pol_header + 1, end_row)
        if table[i][0] == "POD"
    )

    # Parse POL ETA (sail) and Latest Delivery (cutoff)
    pol_data = {}  # (v_idx, pol) -> (cutoff, sail)
    for row in table[pol_header + 1 : pod_row]:
        pol_name = row[0]
        if not pol_name or pol_name not in ACL_POL_MAP:
            continue
        pol = ACL_POL_MAP[pol_name]
        for v_idx in range(len(vessels)):
            sail = parse_acl_date(row[2 * v_idx + 1])
            cutoff = parse_acl_date(row[2 * v_idx + 2])
            pol_data[(v_idx, pol)] = (cutoff, sail)

    # Parse POD arrival dates
    rows = []
    for row in table[pod_row + 1 : end_row]:
        pod_name = row[0]
        if not pod_name or pod_name not in ACL_POD_MAP:
            continue
        pod = ACL_POD_MAP[pod_name]
        for v_idx, (vessel, voyage) in enumerate(vessels):
            arrival = parse_acl_date(row[2 * v_idx + 1])
            if arrival == "NA":
                continue
            for pol in ACL_POL_MAP.values():
                cutoff, sail = pol_data.get((v_idx, pol), ("NA", "NA"))
                rows.append((vessel, voyage, pol, pod, cutoff, sail, arrival))

    return rows


def parse_acl(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        table = pdf.pages[0].extract_tables()[0]

    # Find the two sub-table start rows (rows where col[1] has a vessel name and col[0] is None)
    sub_starts = [
        i for i, row in enumerate(table)
        if row[0] is None and row[1] and row[1].strip().startswith("Grande")
    ]

    rows = []
    for idx, start in enumerate(sub_starts):
        end = sub_starts[idx + 1] if idx + 1 < len(sub_starts) else len(table)
        rows.extend(parse_acl_sub_table(table, start, end))

    return rows


def update_master(sallaum_rows, acl_rows, master_path):
    wb = load_workbook(master_path)
    ws = wb.active

    # Collect vessel names being replaced (case-insensitive)
    new_vessels = {r[0].upper() for r in sallaum_rows + acl_rows}

    # Delete existing rows for those vessels (bottom-up)
    to_delete = [
        i for i in range(2, ws.max_row + 1)
        if ws.cell(i, 1).value and str(ws.cell(i, 1).value).upper() in new_vessels
    ]
    for i in reversed(to_delete):
        ws.delete_rows(i)

    # Append new rows
    for vessel, voyage, pol, pod, cutoff, sail, arrival in sallaum_rows + acl_rows:
        row_data = [
            vessel, voyage, pol, pod,
            cutoff.date() if hasattr(cutoff, "date") else cutoff,
            sail.date() if hasattr(sail, "date") else sail,
            arrival.date() if hasattr(arrival, "date") else arrival,
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for col in (5, 6, 7):
            cell = ws.cell(row_idx, col)
            if hasattr(cell.value, "year"):
                cell.number_format = "MM/DD/YYYY"

    wb.save(master_path)
    return len(to_delete), len(sallaum_rows) + len(acl_rows)


def main():
    if len(sys.argv) < 3:
        print("Usage: python update_schedule.py <sallaum_pdf> <acl_pdf> [master_xlsx]")
        sys.exit(1)

    sallaum_pdf = sys.argv[1]
    acl_pdf = sys.argv[2]
    master_path = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_MASTER

    print(f"Parsing Sallaum PDF: {sallaum_pdf}")
    sallaum_rows = parse_sallaum(sallaum_pdf)
    print(f"  -> {len(sallaum_rows)} rows extracted")

    print(f"Parsing ACL PDF: {acl_pdf}")
    acl_rows = parse_acl(acl_pdf)
    print(f"  -> {len(acl_rows)} rows extracted")

    print(f"Updating master schedule: {master_path}")
    deleted, added = update_master(sallaum_rows, acl_rows, master_path)
    print(f"  -> Deleted {deleted} old rows, added {added} new rows")
    print("Done.")


if __name__ == "__main__":
    main()
